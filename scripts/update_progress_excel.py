# -*- coding: utf-8 -*-
"""重新生成项目进度表 Excel（完整内容，含第16项 P1 需求）"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

wb = openpyxl.Workbook()
thin = Side(style='thin', color='999999')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
bold = Font(bold=True)
header_fill = PatternFill('solid', fgColor='D9E2F3')

# ========== Sheet1 项目总览 ==========
ws1 = wb.active
ws1.title = '项目总览'
rows1 = [
    ['拔蛟窝社区小网格综合治理平台 · 项目进度总览'],
    ['报告日期：2026-08-13    面向客户：拔蛟窝社区两委'],
    [],
    ['整体进度统计', '', '', '三端一屏架构概览'],
    ['需求总数', '16 项', '', '端/屏', '使用者', '形态', '核心场景', '开发状态', '进度'],
    ['已实现', '10 项（63%）', '', '指挥部终端', '总网格负责人、两委', 'Web + BI大屏', '态势看板、工单督办、视频轮巡、应急会商', '已实现', '主体完成'],
    ['部分实现', '1 项（6%）', '', '网格员移动端', '网格组长/组员', '小程序(H5)', '巡查打卡、拍照上报、走访登记、任务接收', '已实现', '主体完成'],
    ['规划中', '5 项（31%）', '', '居民端', '社区居民', '小程序(H5)', '一键上报、进度查询、活动报名、便民报修', '已实现', '主体完成'],
    ['P0 核心需求', '6/6 已实现', '', 'BI大屏', '指挥部', '可视化大屏', '人口密度、事件热力、风险点位、待办工单', '已实现', '主体完成'],
    [],
    ['P0 核心需求（6/6 已实现）'],
    ['GIS 网格可视化', '6大网格+12小网格，AMap多边形边界，深色地图', '', '', '', '已实现'],
    ['人地事物情数据库', '8张新表，人口/房屋/场所/组织全CRUD', '', '', '', '已实现'],
    ['社区指挥部终端', '态势看板、事件闭环、工单处置、三色分级', '', '', '', '已实现'],
    ['工作人员移动端', 'GPS打卡、现场拍照、网格归属判定', '', '', '', '已实现'],
    ['网格巡查', '巡查记录持久化、常规/专项巡查', '', '', '', '已实现'],
    ['事件闭环处置', '三色分级+自动升级+多渠道来源', '', '', '', '已实现'],
]
for r in rows1:
    ws1.append(r)
ws1.merge_cells('A1:I1')
ws1.merge_cells('A2:I2')
ws1.cell(1, 1).font = Font(bold=True, size=16)
for c in ws1[4]:
    c.font = bold
    c.fill = header_fill
for c in ws1[5]:
    c.font = bold
    c.fill = header_fill

# ========== Sheet2 需求功能进度明细 ==========
ws2 = wb.create_sheet('需求功能进度明细')
rows2 = [
    ['编号', '类别', '需求功能', '优先级', '实现状态', '完成度', '实现说明', '待办/备注'],
    [1, '数据库/地图', 'GIS 网格可视化\n6大网格+12小网格三级单元\nAMap多边形边界绘制', 'P0', '已实现', '主体完成', 'AMap深色地图、多边形绘制、点击高亮、树导航+详情面板；打卡以"点"形式展现位置；事件详情点击查看；BI大屏整体展示。微信小程序端已完成地图网格边界多边形+我的网格红色高亮', '—'],
    [2, '数据库', '六大基础库\n实有人口/房屋/场所/组织/事件/政策', 'P0', '已实现', '主体完成', '8张新表创建完成：cmn_grid、cmn_population、cmn_building、cmn_place、cmn_org_member、cmn_patrol_record、cmn_patrol_task、cmn_resident_report；含标签/消防风险/群租房标记；地址+网格选择强制绑定', '—'],
    [3, '数据库', '平台互通与数据标准\n应急/卫健/民政/物业/12345', 'P2', '规划中', '未启动', '需第三方接口对接，涉及政务数据权限审批', '建议二期实施；需确认各平台接口规范与数据标准'],
    [4, '数据库', '效能考核与大数据分析\nKPI报表/月报/热力图/趋势预判', 'P1', '已实现', '基本完成', 'BI看板DashboardView完成：KPI卡片（网格/人口/房屋/场所/组织/事件总数）、三色分级进度条、网格人口排名、高频问题统计；事件上传强制选择类别+网格地址', '四项考核指标口径需最终确认；自动月报模板待完善'],
    [5, '数据库', '电子台账一键导出\n综治/安全生产/爱卫/调解台账', 'P2', '规划中', '数据就绪', '小程序拍照上传作为台账数据来源已就绪；台账导出功能未实现', '需按上级规范补充Excel导出模板'],
    [6, '操作端', '社区指挥部终端\n态势看板/工单督办/视频轮巡/应急会商/三色分级', 'P0', '已实现', '主体完成', 'Web管理端全部页面：态势看板、事件闭环操作、工单处置、审核流程；深蓝政务风格UI；事件三色分级（绿/黄/红）', '一键应急联动（暴雨/火灾/群体性事件）需补充通知链路'],
    [7, '操作端', '网格员移动端\n巡查打卡/拍照上报/走访登记/离线采集', 'P0', '已实现', '主体完成', 'H5页面：GPS定位打卡、网格归属判定、现场拍照、巡查内容记录；巡查历史列表。微信小程序端已完成：底部导航（工作台/地图/巡查/我的）、地图网格边界+我的网格红色高亮、巡查打卡（地图定位/地图选点）、事件上报、手机号验证码登录', '离线采集与弱网回传待验证；操作培训教程待内置'],
    [8, '操作端', '居民端\n一键上报/诉求查询/活动报名/便民报修/随手拍', 'P1', '已实现', '主体完成', 'H5页面：一键上报（投诉/报修/活动/政策/隐患）、自动生成查询码、照片上传；上报历史列表。微信小程序端已完成：随手拍/服务/活动/报修/政策/积分/我的、手机号验证码登录、注册身份选择（居民即时开通/网格员审批）', '积分规则待配置；随手拍自动定位已实现'],
    [9, '工作板块', '网格巡查\n巡查清单/自动排期/未巡预警/自定义到人', 'P0', '已实现', '基本完成', '巡查记录表(cm_patrol_record)+任务表(cm_patrol_task)已创建；支持常规/专项巡查；定位打卡+现场填写', '自动排期（周/月任务）未实现；"到期未巡"系统预警未实现'],
    [10, '工作板块', '停车管理\n车位查询/违停预警/轨迹追踪/精准派单', 'P2', '规划中', '未启动', '依赖视频/地磁等硬件设备', '需求待明确，建议调研后开发'],
    [11, '工作板块', '事件处置闭环\n发现上报→智能派单→现场处置→复核→督办→归档\n三色分级+超期自动升级', 'P0', '已实现', '主体完成', 'biz_event表扩展grid_id/urgency_level/report_source字段；绿(一般)/黄(重点)/红(紧急)三色分级；24h绿→黄、48h黄→红自动升级；多渠道来源（网格员/居民/12345/物业/AI抓拍）', '—'],
    [12, '工作板块', '安全生产/卫生防疫\n九小场所监管/消防隐患/蚊媒管控', 'P1', '部分实现', '约60%', '房屋表含fire_risk_level(LOW/MEDIUM/HIGH)、is_group_rental标记；事件表含urgency_level三色分级', '九小场所"一场一档"动态监管待完善；蚊媒孳生地红黄绿管控未实现'],
    [13, '工作板块', '智慧党建\n联户台账/志愿服务/三会一课/量化考核', 'P2', '规划中', '未启动', '需求已明确但未开始开发', '可作为独立模块开发；党员量化考核需自动数据采集'],
    [14, '设备类', '无人机+视频监控\n每日全域自动巡检/AI识别/九大板块', 'P1', '规划中', '技术预研', '已有drone-api-integration-tasks.md技术预研文档', '需大疆SDK三方对接；AI识别结果自动进入事件闭环待实现'],
    [15, '设备类', '捕蚊器数据量化\n设备监控/数据量化/实物图片', 'P2', '规划中', '未启动', '需求已提出但未开发', '建议以人工录入+照片留证方案先行；IoT数据接入方案待调研'],
    [16, '认证/小程序', '统一认证与微信小程序端\n手机号验证码登录（阿里云短信）\n统一登录入口（验证码⇄账号密码切换）\n注册身份选择（居民即时开通/网格员审批）', 'P1', '已实现', '主体完成', '阿里云短信真实发送验证码（模板SMS_480105081）；统一登录页（手机号验证码⇄账号密码，按角色自动分流网格员/居民）；注册页身份选择（居民直接开通/网格员管理员审批）；微信小程序端全量适配（网格员4tab导航、地图网格多边形、巡查打卡定位选点、事件上报、居民随手拍/服务/报修/积分/政策）；账号手机号唯一绑定', '小程序合法域名需配置 https://drone.kfktec.cn:8443；短信配额40条/号/天；生产建议短信签名/模板走客户主体审核'],
]
for r in rows2:
    ws2.append(r)
for c in ws2[1]:
    c.font = bold
    c.fill = header_fill
    c.border = border
for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row):
    for c in row:
        c.border = border
        c.alignment = Alignment(vertical='top', wrap_text=True)
ws2.column_dimensions['A'].width = 6
ws2.column_dimensions['B'].width = 12
ws2.column_dimensions['C'].width = 32
ws2.column_dimensions['D'].width = 8
ws2.column_dimensions['E'].width = 10
ws2.column_dimensions['F'].width = 10
ws2.column_dimensions['G'].width = 70
ws2.column_dimensions['H'].width = 40
ws2.row_dimensions[18].height = 80

# ========== Sheet3 技术实现清单 ==========
ws3 = wb.create_sheet('技术实现清单')
rows3 = [
    ['数据库表（已创建）'],
    ['表名', '用途', '字段完整性', '索引', '状态'],
    ['cmn_grid', '网格层级（6大+12小）', '完整', 'parent_id, level, status', '✅'],
    ['cmn_population', '实有人口（含标签/类型）', '完整', 'grid_id, household_type, status', '✅'],
    ['cmn_building', '房屋/出租屋（含消防/群租标记）', '完整', 'grid_id, fire_risk_level', '✅'],
    ['cmn_place', '场所资源（九小场所等）', '完整', 'grid_id, place_type', '✅'],
    ['cmn_org_member', '组织力量（两委/网格员/志愿者）', '完整', 'grid_id, member_type', '✅'],
    ['cmn_patrol_record', '巡查打卡记录', '完整', 'grid_id, user_id', '✅'],
    ['cmn_patrol_task', '巡查任务', '完整', 'grid_id, user_id, status, planned_date', '✅'],
    ['cmn_resident_report', '居民诉求/随手拍', '完整', 'grid_id, status, query_code', '✅'],
    ['biz_event（扩展）', '事件闭环（+grid_id/urgency/report_source）', '完整', 'grid_id, urgency_level', '✅'],
    [],
    ['Web 管理端页面'],
    ['页面', '路由', '模板', '功能完整性', '状态'],
    ['GIS网格可视化', '/community/grid', '自定义(AMap)', '地图+多边形+树导航', '✅'],
    ['实有人口管理', '/community/population', 'WebListPageTemplate', '筛选+CRUD', '✅'],
    ['房屋管理', '/community/buildings', 'WebListPageTemplate', '筛选+CRUD+消防标记', '✅'],
    ['场所管理', '/community/places', 'WebListPageTemplate', '类型筛选+CRUD', '✅'],
    ['组织管理', '/community/org-members', 'WebListPageTemplate', '类型筛选+CRUD', '✅'],
    ['BI态势看板', '/community/dashboard', '自定义', 'KPI+三色分级+排名', '✅'],
    ['事件列表', '/events', '已有扩展', '三色分级标签', '✅'],
    ['工单列表', '/work-orders', '已有', '复用', '✅'],
    ['智能派单规则', '/dispatch-rules', '自定义', '规则配置+一键派单', '✅'],
    ['网格管理', '/community/grid-manage', '自定义', '网格调整+数量管理', '✅'],
    [],
    ['微信小程序端（新增）'],
    ['页面', '路径', '功能', '状态', ''],
    ['统一登录页', 'pages/role-select/index', '手机号验证码⇄账号密码、身份自动分流', '✅', ''],
    ['统一注册页', 'pages/register/index', '身份选择（居民即时/网格员审批）', '✅', ''],
    ['网格员工作台', 'pages/workbench/index', '待办统计+快捷入口+4tab导航', '✅', ''],
    ['移动GIS地图', 'pages/map/index', '网格边界多边形+事件点+我的网格高亮', '✅', ''],
    ['巡查打卡', 'pages/patrol/checkin', '地图定位/选点+拍照+网格选择', '✅', ''],
    ['事件上报', 'pages/event/report', '类型+标题+描述+定位+照片', '✅', ''],
    ['居民随手拍', 'pages/resident/report/index', '类型+描述+照片+定位选点', '✅', ''],
    ['居民服务/报修/积分/政策', 'pages/resident/*', '活动报名、便民报修、积分、政策', '✅', ''],
]
for r in rows3:
    ws3.append(r)
for c in ws3[1]:
    c.font = bold
    c.fill = header_fill
for c in ws3[13]:
    c.font = bold
    c.fill = header_fill
for c in ws3[26]:
    c.font = bold
    c.fill = header_fill

out = 'C:/Users/Administrator/Desktop/拔蛟窝社区小网格治理平台_项目进度表_更新.xlsx'
wb.save(out)
print('已重新生成:', out)
